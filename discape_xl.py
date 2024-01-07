from openpyxl import Workbook
from openpyxl import load_workbook

file = 'file.xlsx'

wb = load_workbook(filename=file)

CHAR_NAME_COL = 0
CHAR_USER_COL = 1
CHAR_ROOM_COL = 2
CHAR_PATH_COL = 3
CHAR_HAND_COL = 4

def check_valid_format() -> bool:
    """Checks if the loaded file has the correct format."""
    wb = load_workbook(filename=file)
    if len(wb.sheetnames) < 4:
        return False
    return True

def get_column_values(sheet_name: str, column: int) -> list[str]:
    """Returns a list with the values of the specified column. Includes empty cells."""
    ws = wb[sheet_name]
    return [row[column].value for row in ws]

def get_row_values(sheet_name: str, row: int) -> list[str]:
    """Returns a list with the values of the specified row. Includes empty cells."""
    ws = wb[sheet_name]
    return [cell.value for cell in ws[row]]

def remove_row(sheet_name: str, row: int):
    """Removes the specified row from the specified sheet."""
    ws = wb[sheet_name]
    ws.delete_rows(row)
    wb.save(file)

def append_row(sheet_name: str, row_values: list):
    """Adds a new row at the end of the specified sheet."""
    ws = wb[sheet_name]
    new_row = ws.max_row + 1
    for i in range(len(row_values)):
        ws.cell(row=new_row, column=i + 1, value=row_values[i])
    wb.save(file)

def get_stat(player: str, stat: str) -> int:
    """Returns the value of the specified stat for the specified player."""
    sheet_name = 'Personajes'
    ws = wb[sheet_name]
    player_column = CHAR_USER_COL

    # Get the name of the stats and their column number
    stat_col = {}
    for i in range(5, ws.max_column + 1):
        stat_col[ws.cell(row=1, column=i).value] = i

    player_row = get_column_values(sheet_name, player_column).index(player) + 1
    return ws.cell(row=player_row, column=stat_col[stat]).value

def get_player_location(player: str) -> tuple[str, str]:
    """Returns the location of the specified player as a tuple containing the name of the room and the path taken."""
    ws = wb['Personajes']
    sheet_name = 'Personajes'
    player_column = CHAR_USER_COL
    location_column = CHAR_ROOM_COL
    path_column = CHAR_PATH_COL
    player_column_values = get_column_values(sheet_name, player_column)
    player_row = player_column_values.index(player)
    return ws.cell(row=player_row + 1, column=location_column + 1).value, ws.cell(row=player_row, column=path_column + 1).value

def update_player_location(player: str, location: str, path: str) -> None:
    """Updates the location of the specified player."""
    ws = wb['Personajes']
    sheet_name = 'Personajes'
    player_column = CHAR_USER_COL
    location_column = CHAR_ROOM_COL
    path_column = CHAR_PATH_COL
    player_row = get_column_values(sheet_name, player_column).index(player) + 1
    ws.cell(row=player_row, column=location_column + 1, value=location)
    ws.cell(row=player_row, column=path_column + 1, value=path)
    wb.save(file)

def get_zones(player: str) -> list[str]:
    """Returns a list with the names of the zones."""
    (room, path) = get_player_location(player)
    sheet_name = room
    ws = wb[sheet_name]
    depth_col = 2
    key_col = 4

    zones = []
    for row in ws:
        if row[depth_col].value == path and row[key_col].value == None:
            zones.append(row[0].value)
    return zones

def unlock_zone(player: str, item: str) -> None:
    """Unlocks the zone with the specified item."""
    (room, path) = get_player_location(player)
    sheet_name = room
    ws = wb[sheet_name]
    depth_col = 2

    for row in ws:
        if len(row) >= 5:
            if row[depth_col] == path and row[4] == item:
                row[4] = ''
                wb.save(file)
                break

def add_item(new_data: list[str]) -> None:
    """Adds a new item to the inventory.
    The new item must be a list with the following format: [name, description]."""
    append_row('Inventario', new_data)

def get_inventory() -> list[str]:
    """Returns a list with the names of the items in the inventory."""
    ws = wb['Inventario']
    return [row[0].value for row in ws]

def combine(item1: str, item2: str):  # TODO: This is not finished
    """Combines the two specified items and returns the resulting item."""
    ws = wb['Combinaciones']
    for row in ws:
        if row[0].value == item1 and row[1].value == item2:
            return row[2].value
        elif row[0].value == item2 and row[1].value == item1:
            return row[2].value

def unlock_item(room: str, item: str):
    """Unlocks the specified item in the specified room."""
    ws = wb[room]
    for row in ws:
        if item in row:
            add_item([row[0:2]])
            remove_row(room, row.index(item))
    wb.save(file)

def take_path(player: str, choice: str):
     """Moves the player to the specified path or adds the item to the inventory.
    Returns the description of the new room or the description of the item."""
     (room, path) = get_player_location(player)
     ws = wb[room]
     depth_col = 2   # FIXME: This could be found automatically some other way, probably
     if choice == '↩️ Volver':
        update_player_location(player, room, path[:-1])
        return "Volviste al lugar anterior."
     for row in ws:
         if row[depth_col].value == path and row[0].value == choice:
             if row[3] != 'Objeto':
                 update_player_location(player, room, path + row[3].value)
             else:
                unlock_item(room, choice)
             wb.save(file)
             return row[1].value

# print(get_column_values('Personajes', 1))
# data = get_row_values('Personajes', 2)
# append_row('Personajes', data)
# print(get_column_values('Personajes', 1))
# remove_row('Personajes', 4)
# print(get_column_values('Personajes', 1))
# print(get_stat('reimeko', 'Suerte'))
# update_player_location('reimeko', 'Sala de prueba', 'A')
print(get_zones('reimeko'))
