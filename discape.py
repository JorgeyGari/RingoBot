from openpyxl import load_workbook
import puzzles
import discord

file = "file.xlsx"
wb = load_workbook(filename=file)

CHAR_NAME_COL = 0
CHAR_USER_COL = 1
CHAR_ROOM_COL = 2
CHAR_PATH_COL = 3
CHAR_HAND_COL = 4


class PaginationView(discord.ui.View):
    def __init__(self, embeds):
        super().__init__()
        self.embeds = embeds
        self.current = 0

    async def update_page(self, interaction: discord.Interaction):
        self.embeds[self.current].set_footer(
            text=f"Página {self.current + 1} de {len(self.embeds)}"
        )
        await interaction.response.edit_message(embed=self.embeds[self.current])

    @discord.ui.button(label="⬅️", style=discord.ButtonStyle.blurple)
    async def previous_page(
        self, _: discord.ui.Button, interaction: discord.Interaction
    ):
        self.current = (self.current - 1) % len(self.embeds)
        await self.update_page(interaction)

    @discord.ui.button(label="➡️", style=discord.ButtonStyle.blurple)
    async def next_page(self, _: discord.ui.Button, interaction: discord.Interaction):
        self.current = (self.current + 1) % len(self.embeds)
        await self.update_page(interaction)


def get_column_values(sheet_name: str, column: int) -> list[str]:
    """Returns a list with the values of the specified column. Includes empty cells."""
    ws = wb[sheet_name]
    return [row[column].value for row in ws]


def get_row_values(sheet_name: str, row: int) -> list[str]:
    """Returns a list with the values of the specified row. Includes empty cells."""
    ws = wb[sheet_name]
    return [cell.value for cell in ws[row]]


def remove_row(sheet_name: str, row: int) -> None:
    """Removes the specified row from the specified sheet."""
    ws = wb[sheet_name]
    ws.delete_rows(row)
    wb.save(file)


def append_row(sheet_name: str, row_values: list):
    """Adds a new row at the end of the specified sheet."""
    ws = wb[sheet_name]
    new_row = ws.max_row + 1
    for i in range(len(row_values)):
        ws.cell(row=new_row, column=i + 1, value=row_values[i])
    wb.save(file)


def get_stat(player: str, stat: str) -> int:
    """Returns the value of the specified stat for the specified player."""
    sheet_name = "Personajes"
    ws = wb[sheet_name]
    player_column = CHAR_USER_COL

    # Get the name of the stats and their column number
    stat_col = {}
    for i in range(5, ws.max_column + 1):
        stat_col[ws.cell(row=1, column=i).value] = i

    player_row = get_column_values(sheet_name, player_column).index(player) + 1
    return ws.cell(row=player_row, column=stat_col[stat]).value


def get_player_location(player: str) -> tuple[str, str]:
    """Returns the location of the specified player as a tuple containing the name of the room and the path taken."""
    ws = wb["Personajes"]
    sheet_name = "Personajes"
    player_column = CHAR_USER_COL
    location_column = CHAR_ROOM_COL
    path_column = CHAR_PATH_COL
    player_column_values = get_column_values(sheet_name, player_column)
    player_row = player_column_values.index(player)
    return (
        ws.cell(row=player_row + 1, column=location_column + 1).value,
        ws.cell(row=player_row + 1, column=path_column + 1).value,
    )


def get_player_hand(player: str) -> str:
    """Returns the item in the player's hand."""
    ws = wb["Personajes"]
    sheet_name = "Personajes"
    player_column = CHAR_USER_COL
    hand_column = CHAR_HAND_COL
    player_column_values = get_column_values(sheet_name, player_column)
    player_row = player_column_values.index(player)
    return ws.cell(row=player_row + 1, column=hand_column + 1).value


def update_player_location(player: str, location: str, path: str | None) -> None:
    """Updates the location of the specified player."""
    ws = wb["Personajes"]
    sheet_name = "Personajes"
    player_column = CHAR_USER_COL
    location_column = CHAR_ROOM_COL
    path_column = CHAR_PATH_COL
    player_row = get_column_values(sheet_name, player_column).index(player)
    ws.cell(row=player_row + 1, column=location_column + 1, value=location)
    if path == "":
        ws.cell(row=player_row + 1, column=path_column + 1).value = None
    else:
        ws.cell(row=player_row + 1, column=path_column + 1, value=path)
    wb.save(file)


def get_zones(player: str) -> list[str]:
    """Returns a list with the names of the zones."""
    (room, path) = get_player_location(player)
    sheet_name = room
    ws = wb[sheet_name]
    depth_col = 2
    key_col = 4

    zones = []
    for row in ws:
        if row[depth_col].value == path and row[key_col].value == None:
            zones.append(row[0].value)
    return zones


def unlock_zone(player: str, item: str) -> str | None:
    """Unlocks the zone with the specified item."""
    (room, path) = get_player_location(player)
    sheet_name = room
    ws = wb[sheet_name]
    depth_col = 2
    key_col = 4
    act_col = 5

    for row in ws:
        before = row[depth_col].value[:-1] if row[depth_col].value else None
        if before == "":
            before = None
        if row[key_col].value == item and path == before:
            row[key_col].value = None
            update_player_location(player, room, row[depth_col].value)
            wb.save(file)
            return row[act_col].value


def add_item(new_data: list[str]) -> None:
    """Adds a new item to the inventory.
    The new item must be a list with the following format: [name, description, room]."""
    append_row("Inventario", new_data)


def remove_item(item: str, room: str) -> None:
    """Removes the specified item from the inventory."""
    ws = wb["Inventario"]
    for row in ws:
        if row[0].value == item and row[2].value == room:
            remove_row("Inventario", row[0].row)
            break
    wb.save(file)


def get_inventory_dict(room) -> dict[str, str]:
    """Returns a dictionary with the item's name as the key and the item's description as the value."""
    ws = wb["Inventario"]
    return {
        row[0].value: row[1].value
        for row in ws.iter_rows(min_row=2)
        if row[0].value is not None
        and row[1].value is not None
        and row[2].value == room
    }


def get_inventory_names(room) -> list[str]:
    """Returns a list with the names of the items in the inventory of a specified room."""
    ws = wb["Inventario"]
    inventory = []
    for row in ws:
        if row[2].value == room:
            inventory.append(row[0].value)
    return inventory


def equip(player, item) -> str:
    """Equips the player with the specified item."""
    ws = wb["Personajes"]
    sheet_name = "Personajes"
    player_column = CHAR_USER_COL
    hand_column = CHAR_HAND_COL
    room = get_player_location(player)[0]
    player_row = get_column_values(sheet_name, player_column).index(player)
    if item in get_inventory_names(room):
        ws.cell(row=player_row + 1, column=hand_column + 1, value=item)
        wb.save(file)
        return f"Equipaste: {item}."
    else:
        return f"No tienes eso."


def combine(item1: str, item2: str, room) -> str:
    """Combines the two specified items and returns the resulting item.
    Returns None if the combination is not valid."""
    if item1 not in get_inventory_names(room) or item2 not in get_inventory_names(room):
        return "No tienes esos objetos."

    ws = wb["Combinaciones"]
    item1_col = 0
    item2_col = 1
    result_col = 2
    desc_col = 3
    room_col = 4
    for row in ws:
        if (
            row[item1_col].value == item1
            and row[item2_col].value == item2
            and row[room_col].value == room
        ):
            add_item([row[result_col].value, row[desc_col].value, room])
            remove_item(item1, room)
            remove_item(item2, room)
            return "Has combinado los objetos y has obtenido un nuevo objeto."
        elif (
            row[item1_col].value == item2
            and row[item2_col].value == item1
            and row[room_col].value == room
        ):
            add_item([row[result_col].value, row[desc_col].value, room])
            remove_item(item1, room)
            remove_item(item2, room)
            return "Has combinado los objetos y has obtenido un nuevo objeto."
    return "No puedes combinar esos objetos."


def unlock_item(room: str, item: str) -> str:
    """Unlocks the specified item in the specified room."""
    room = wb[room]
    for attraction in room:
        if attraction[0].value == item:
            add_item([item, attraction[1].value, room.title])
            remove_row(room.title, attraction[0].row)
            return f"Has obtenido un nuevo objeto: {item}."
    wb.save(file)


def get_path_from_choice(player: str, choice: str) -> str | None:
    """Returns the key of the specified choice."""
    (room, _) = get_player_location(player)
    ws = wb[room]
    name_col = 0
    path_col = 3

    for row in ws:
        if row[name_col].value == choice:
            return row[path_col].value


def get_key_from_path(player: str, path: str) -> str | None:
    """Returns the key of the specified path."""
    (room, _) = get_player_location(player)
    ws = wb[room]
    depth_col = 2
    key_col = 4

    for row in ws:
        if row[depth_col].value == path:
            return row[key_col].value


def take_path(player: str, choice: str) -> str:
    """Moves the player to the specified path or adds the item to the inventory.
    Returns the description of the new room or the description of the item."""
    (room, path) = get_player_location(player)

    if room is None:
        return "No estás en ninguna sala."

    ws = wb[room]
    name_col = 0
    depth_col = 2
    path_col = 3
    hand = get_player_hand(player)

    if choice == "↩️ Volver":
        update_player_location(player, room, path[:-1])
        wb.save(file)
        return "Volviste al lugar anterior."

    if hand:
        if path is not None:
            key = get_key_from_path(player, path + get_path_from_choice(player, choice))
        else:
            key = get_key_from_path(player, get_path_from_choice(player, choice))
        if key == hand:
            result = unlock_zone(player, hand)
            if result:
                return result

    for row in ws:
        if (
            row[depth_col].value == path and row[name_col].value == choice
        ):  # Found the choice
            if row[path_col].value == "Final":
                return escaped(room)
            if row[path_col].value == "Objeto":
                return unlock_item(room, choice)
            if row[path_col].value.startswith("Puzle"):
                puzzle_name = row[path_col].value.split(" ")[1]
                return puzzles.solve_puzzle(puzzle_name)
            else:
                if path and row[path_col].value:
                    update_player_location(player, room, path + row[path_col].value)
                elif row[path_col].value:
                    update_player_location(player, room, row[path_col].value)
            wb.save(file)
            return row[1].value
    return "No puedes ir por ahí."


def escaped(room: str) -> str:
    """Returns the description of the room and removes the player from the game."""
    ws = wb["Personajes"]
    for row in ws:
        if row[CHAR_ROOM_COL].value == room:
            ws.cell(row=row[CHAR_NAME_COL].row, column=CHAR_ROOM_COL + 1).value = None
    wb.save(file)
    return "**Has escapado.**"


def join_room(player: str, room: str) -> str:
    """Adds the player to the specified room and returns the description of the room."""
    ws = wb["Personajes"]
    for row in ws:
        if row[CHAR_USER_COL].value == player:
            ws.cell(row=row[CHAR_NAME_COL].row, column=CHAR_ROOM_COL + 1).value = room
    wb.save(file)
    return f"Te has unido a la sala: {room}."


# # TESTS
# print(get_column_values("Personajes", 1))
# data = get_row_values("Personajes", 2)
# append_row("Personajes", data)
# print(get_column_values("Personajes", 1))
# remove_row("Personajes", 4)
# print(get_column_values("Personajes", 1))
# print(get_stat("reimeko", "Suerte"))
# update_player_location("reimeko", "Sala de prueba", "A")
# print(get_zones("jorgeygari"))
# print(combine("Destornillador", "Caja"))
# print(get_inventory())
# add_item(["Destornillador", "Un destornillador."])
# print(get_inventory())
# remove_row("Inventario", 2)
# print(get_inventory())
# unlock_item("Sala de prueba", "Caja")
# print(get_player_hand("reimeko"))
# print(get_player_hand("jorgeygari"))
# print("fail")
# unlock_zone("reimeko", "Antorcha")
# update_player_location("reimeko", "Sala de prueba", "A")
# print("success")
# unlock_zone("reimeko", "Antorcha")
# print(get_inventory_dict())
# print(combine("Destornillador", "Caja"))
# print(get_path_from_choice("jorgeygari", "Puerta"))
# print(get_key_from_path("jorgeygari", "BF"))
