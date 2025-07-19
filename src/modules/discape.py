"""
Discape module for escape room simulation functionality.
"""

import os
import logging
from openpyxl import load_workbook
import discord
from typing import List, Tuple, Dict, Optional

from utils.config import config

logger = logging.getLogger(__name__)


class PaginationView(discord.ui.View):
    """Discord UI view for paginated content."""

    def __init__(self, embeds):
        super().__init__()
        self.embeds = embeds
        self.current = 0

    async def update_page(self, interaction: discord.Interaction):
        """Update the current page display."""
        self.embeds[self.current].set_footer(
            text=f"Página {self.current + 1} de {len(self.embeds)}"
        )
        await interaction.response.edit_message(embed=self.embeds[self.current])

    @discord.ui.button(label="⬅️", style=discord.ButtonStyle.blurple)
    async def previous_page(
        self, _: discord.ui.Button, interaction: discord.Interaction
    ):
        """Go to previous page."""
        self.current = (self.current - 1) % len(self.embeds)
        await self.update_page(interaction)

    @discord.ui.button(label="➡️", style=discord.ButtonStyle.blurple)
    async def next_page(self, _: discord.ui.Button, interaction: discord.Interaction):
        """Go to next page."""
        self.current = (self.current + 1) % len(self.embeds)
        await self.update_page(interaction)


class DiscapeModule:
    """Handles Discape escape room functionality."""

    # Column indices for character sheet
    CHAR_NAME_COL = 0
    CHAR_USER_COL = 1
    CHAR_ROOM_COL = 2
    CHAR_PATH_COL = 3
    CHAR_HAND_COL = 4

    def __init__(self):
        """Initialize the Discape module."""
        self.wb = None
        self._load_workbook()

    def _load_workbook(self):
        """Load the Excel workbook for Discape data."""
        try:
            if os.path.exists(config.DISCAPE_FILE):
                self.wb = load_workbook(filename=config.DISCAPE_FILE)
                logger.info("Discape workbook loaded successfully")
            else:
                logger.warning(f"Discape file not found: {config.DISCAPE_FILE}")
        except Exception as e:
            logger.error(f"Error loading Discape workbook: {e}")

    def get_column_values(self, sheet_name: str, column: int) -> List[str]:
        """Get all values from a specific column in a sheet."""
        if not self.wb:
            return []

        try:
            ws = self.wb[sheet_name]
            return [row[column].value for row in ws]
        except Exception as e:
            logger.error(f"Error getting column values: {e}")
            return []

    def get_row_values(self, sheet_name: str, row: int) -> List[str]:
        """Get all values from a specific row in a sheet."""
        if not self.wb:
            return []

        try:
            ws = self.wb[sheet_name]
            return [cell.value for cell in ws[row]]
        except Exception as e:
            logger.error(f"Error getting row values: {e}")
            return []

    def get_player_location(self, player: str) -> Tuple[str, str]:
        """Get the current location of a player."""
        if not self.wb:
            return (None, None)

        try:
            ws = self.wb["Personajes"]
            player_column_values = self.get_column_values(
                "Personajes", self.CHAR_USER_COL
            )
            player_row = player_column_values.index(player)

            room = ws.cell(row=player_row + 1, column=self.CHAR_ROOM_COL + 1).value
            path = ws.cell(row=player_row + 1, column=self.CHAR_PATH_COL + 1).value

            return (room, path)
        except (ValueError, Exception) as e:
            logger.error(f"Error getting player location: {e}")
            return (None, None)

    def get_player_hand(self, player: str) -> Optional[str]:
        """Get the item in player's hand."""
        if not self.wb:
            return None

        try:
            ws = self.wb["Personajes"]
            player_column_values = self.get_column_values(
                "Personajes", self.CHAR_USER_COL
            )
            player_row = player_column_values.index(player)

            return ws.cell(row=player_row + 1, column=self.CHAR_HAND_COL + 1).value
        except (ValueError, Exception) as e:
            logger.error(f"Error getting player hand: {e}")
            return None

    def update_player_location(
        self, player: str, location: str, path: Optional[str]
    ) -> None:
        """Update player's location."""
        if not self.wb:
            return

        try:
            ws = self.wb["Personajes"]
            player_row = self.get_column_values("Personajes", self.CHAR_USER_COL).index(
                player
            )

            ws.cell(row=player_row + 1, column=self.CHAR_ROOM_COL + 1, value=location)

            if path == "":
                ws.cell(row=player_row + 1, column=self.CHAR_PATH_COL + 1).value = None
            else:
                ws.cell(row=player_row + 1, column=self.CHAR_PATH_COL + 1, value=path)

            self.wb.save(config.DISCAPE_FILE)
            logger.info(f"Updated player {player} location to {location}, path: {path}")
        except Exception as e:
            logger.error(f"Error updating player location: {e}")

    def get_inventory_dict(self, room: str) -> Dict[str, str]:
        """Get inventory items for a room as a dictionary."""
        if not self.wb:
            return {}

        try:
            ws = self.wb["Inventario"]
            inventory = {}

            for row in ws.iter_rows(min_row=2):
                if (
                    row[0].value is not None
                    and row[1].value is not None
                    and row[2].value == room
                ):
                    inventory[row[0].value] = row[1].value

            return inventory
        except Exception as e:
            logger.error(f"Error getting inventory: {e}")
            return {}

    def join_room(self, player: str, room: str) -> str:
        """Add player to specified room."""
        if not self.wb:
            return "Error: No se pudo cargar el archivo de datos."

        try:
            ws = self.wb["Personajes"]

            for row in ws:
                if row[self.CHAR_USER_COL].value == player:
                    ws.cell(
                        row=row[self.CHAR_NAME_COL].row,
                        column=self.CHAR_ROOM_COL + 1,
                        value=room,
                    )
                    break

            self.wb.save(config.DISCAPE_FILE)
            logger.info(f"Player {player} joined room {room}")
            return f"Te has unido a la sala: {room}."
        except Exception as e:
            logger.error(f"Error joining room: {e}")
            return "Error al unirse a la sala."

    def get_stat(self, player: str, stat: str) -> int:
        """Get the value of the specified stat for the specified player."""
        if not self.wb:
            return 0

        try:
            ws = self.wb["Personajes"]
            player_column_values = self.get_column_values(
                "Personajes", self.CHAR_USER_COL
            )

            # Get the stat column numbers
            stat_col = {}
            for i in range(5, ws.max_column + 1):
                stat_name = ws.cell(row=1, column=i).value
                if stat_name:
                    stat_col[stat_name] = i

            if stat not in stat_col:
                return 0

            player_row = player_column_values.index(player) + 1
            return ws.cell(row=player_row, column=stat_col[stat]).value or 0
        except (ValueError, Exception) as e:
            logger.error(f"Error getting stat: {e}")
            return 0

    def get_zones(self, player: str) -> List[str]:
        """Get list of available zones for a player."""
        if not self.wb:
            return []

        try:
            room, path = self.get_player_location(player)
            if not room:
                return []

            ws = self.wb[room]
            zones = []

            for row in ws:
                if (
                    row[2].value == path and row[4].value is None
                ):  # depth_col and key_col
                    zones.append(row[0].value)

            return zones
        except Exception as e:
            logger.error(f"Error getting zones: {e}")
            return []

    def get_inventory_names(self, room: str) -> List[str]:
        """Get list of item names in a room's inventory."""
        if not self.wb:
            return []

        try:
            ws = self.wb["Inventario"]
            inventory = []

            for row in ws:
                if row[2].value == room and row[0].value:
                    inventory.append(row[0].value)

            return inventory
        except Exception as e:
            logger.error(f"Error getting inventory names: {e}")
            return []

    def add_item(self, new_data: List[str]) -> None:
        """Add a new item to the inventory."""
        if not self.wb:
            return

        try:
            ws = self.wb["Inventario"]
            new_row = ws.max_row + 1
            for i, value in enumerate(new_data):
                ws.cell(row=new_row, column=i + 1, value=value)
            self.wb.save(config.DISCAPE_FILE)
        except Exception as e:
            logger.error(f"Error adding item: {e}")

    def remove_item(self, item: str, room: str) -> None:
        """Remove an item from the inventory."""
        if not self.wb:
            return

        try:
            ws = self.wb["Inventario"]
            for row in ws:
                if row[0].value == item and row[2].value == room:
                    ws.delete_rows(row[0].row)
                    break
            self.wb.save(config.DISCAPE_FILE)
        except Exception as e:
            logger.error(f"Error removing item: {e}")

    def equip(self, player: str, item: str) -> str:
        """Equip a player with the specified item."""
        if not self.wb:
            return "Error: No se pudo cargar el archivo de datos."

        try:
            room = self.get_player_location(player)[0]
            if item in self.get_inventory_names(room):
                ws = self.wb["Personajes"]
                player_column_values = self.get_column_values(
                    "Personajes", self.CHAR_USER_COL
                )
                player_row = player_column_values.index(player)

                ws.cell(row=player_row + 1, column=self.CHAR_HAND_COL + 1, value=item)
                self.wb.save(config.DISCAPE_FILE)
                return f"Equipaste: {item}."
            else:
                return "No tienes eso."
        except Exception as e:
            logger.error(f"Error equipping item: {e}")
            return "Error al equipar el objeto."

    def combine(self, item1: str, item2: str, room: str) -> str:
        """Combine two items and return the result."""
        if not self.wb:
            return "Error: No se pudo cargar el archivo de datos."

        try:
            if item1 not in self.get_inventory_names(
                room
            ) or item2 not in self.get_inventory_names(room):
                return "No tienes esos objetos."

            ws = self.wb["Combinaciones"]

            for row in ws:
                # Check both combinations (item1+item2 and item2+item1)
                if (
                    (row[0].value == item1 and row[1].value == item2)
                    or (row[0].value == item2 and row[1].value == item1)
                ) and row[4].value == room:

                    # Add the resulting item
                    self.add_item([row[2].value, row[3].value, room])
                    # Remove the original items
                    self.remove_item(item1, room)
                    self.remove_item(item2, room)
                    return "Has combinado los objetos y has obtenido un nuevo objeto."

            return "No puedes combinar esos objetos."
        except Exception as e:
            logger.error(f"Error combining items: {e}")
            return "Error al combinar objetos."

    def take_path(self, player: str, choice: str) -> str:
        """Navigate through the escape room or interact with objects."""
        if not self.wb:
            return "Error: No se pudo cargar el archivo de datos."

        try:
            room, path = self.get_player_location(player)

            if room is None:
                return "No estás en ninguna sala."

            ws = self.wb[room]
            hand = self.get_player_hand(player)

            if choice == "↩️ Volver":
                new_path = path[:-1] if path else None
                self.update_player_location(player, room, new_path)
                return "Volviste al lugar anterior."

            # Check if player has item that can unlock something
            if hand:
                if path:
                    target_path = path + self.get_path_from_choice(player, choice)
                else:
                    target_path = self.get_path_from_choice(player, choice)

                key = self.get_key_from_path(player, target_path)
                if key == hand:
                    result = self.unlock_zone(player, hand)
                    if result:
                        return result

            # Process the choice
            for row in ws:
                if (
                    row[2].value == path and row[0].value == choice
                ):  # depth_col and name_col
                    if row[3].value == "Final":  # path_col
                        return self.escaped(room)
                    elif row[3].value == "Objeto":
                        return self.unlock_item(room, choice)
                    elif row[3].value and row[3].value.startswith("Puzle"):
                        puzzle_name = row[3].value.split(" ")[1]
                        # Would need to import puzzles module here
                        return f"Puzzle: {puzzle_name} (not implemented yet)"
                    else:
                        # Navigate to new location
                        if path and row[3].value:
                            new_path = path + row[3].value
                        elif row[3].value:
                            new_path = row[3].value
                        else:
                            new_path = path

                        self.update_player_location(player, room, new_path)
                        return row[1].value  # description

            return "No puedes ir por ahí."
        except Exception as e:
            logger.error(f"Error taking path: {e}")
            return "Error al procesar la acción."

    def get_path_from_choice(self, player: str, choice: str) -> str:
        """Get the path value for a choice."""
        if not self.wb:
            return ""

        try:
            room, _ = self.get_player_location(player)
            ws = self.wb[room]

            for row in ws:
                if row[0].value == choice:  # name_col
                    return row[3].value or ""  # path_col
            return ""
        except Exception as e:
            logger.error(f"Error getting path from choice: {e}")
            return ""

    def get_key_from_path(self, player: str, path: str) -> str:
        """Get the key required for a path."""
        if not self.wb:
            return ""

        try:
            room, _ = self.get_player_location(player)
            ws = self.wb[room]

            for row in ws:
                if row[2].value == path:  # depth_col
                    return row[4].value or ""  # key_col
            return ""
        except Exception as e:
            logger.error(f"Error getting key from path: {e}")
            return ""

    def unlock_zone(self, player: str, item: str) -> str:
        """Unlock a zone with an item."""
        if not self.wb:
            return ""

        try:
            room, path = self.get_player_location(player)
            ws = self.wb[room]

            for row in ws:
                before = row[2].value[:-1] if row[2].value else None  # depth_col
                if before == "":
                    before = None

                if row[4].value == item and path == before:  # key_col
                    row[4].value = None
                    self.update_player_location(player, room, row[2].value)
                    self.wb.save(config.DISCAPE_FILE)
                    return row[5].value or ""  # action_col
            return ""
        except Exception as e:
            logger.error(f"Error unlocking zone: {e}")
            return ""

    def unlock_item(self, room: str, item: str) -> str:
        """Unlock an item in the room."""
        if not self.wb:
            return "Error: No se pudo cargar el archivo de datos."

        try:
            ws = self.wb[room]
            for row in ws:
                if row[0].value == item:  # name_col
                    self.add_item([item, row[1].value, room])
                    ws.delete_rows(row[0].row)
                    self.wb.save(config.DISCAPE_FILE)
                    return f"Has obtenido un nuevo objeto: {item}."
            return "No se pudo obtener el objeto."
        except Exception as e:
            logger.error(f"Error unlocking item: {e}")
            return "Error al obtener el objeto."

    def escaped(self, room: str) -> str:
        """Handle player escaping from the room."""
        if not self.wb:
            return "Error: No se pudo cargar el archivo de datos."

        try:
            ws = self.wb["Personajes"]
            for row in ws:
                if row[self.CHAR_ROOM_COL].value == room:
                    ws.cell(
                        row=row[self.CHAR_NAME_COL].row, column=self.CHAR_ROOM_COL + 1
                    ).value = None

            self.wb.save(config.DISCAPE_FILE)
            return "**Has escapado.**"
        except Exception as e:
            logger.error(f"Error handling escape: {e}")
            return "Error al escapar."

    def get_investigation_options(self, player: str) -> List[str]:
        """Get available investigation options for a player."""
        try:
            room, path = self.get_player_location(player)
            if room:
                zones = self.get_zones(player)
                # Add "go back" option if player has a path
                if room and path:
                    zones.append("↩️ Volver")
                return zones
            else:
                return []
        except Exception as e:
            logger.error(f"Error getting investigation options: {e}")
            return []

    def get_equipable_items_for_player(self, player: str) -> List[str]:
        """Get list of items that a player can equip."""
        try:
            room, _ = self.get_player_location(player)
            if room:
                return self.get_inventory_names(room)
            else:
                return []
        except Exception as e:
            logger.error(f"Error getting equipable items: {e}")
            return []

    # Command handler methods
    async def handle_start_command(self, ctx, archivo):
        """Handle the start escape room command."""
        try:
            await ctx.defer()

            # Save the uploaded file
            await archivo.save(config.DISCAPE_FILE)

            # Reload the workbook
            self._load_workbook()

            await ctx.followup.send("Archivo cargado.")

        except Exception as e:
            logger.error(f"Error in start command: {e}")
            await ctx.followup.send("Error al cargar el archivo.", ephemeral=True)

    async def handle_stat_roll_command(self, ctx, característica: str):
        """Handle the stat roll command."""
        try:
            await ctx.defer()

            player = ctx.user.name
            bonus = self.get_stat(player, característica)

            import random
            roll = random.randint(1, 20)
            result = roll + bonus

            await ctx.followup.send(
                f"¡Has usado {característica}!\n"
                f"`[{roll}] + {bonus}`\n"
                f"**Total:** {result}"
            )

        except Exception as e:
            logger.error(f"Error in stat roll command: {e}")
            await ctx.followup.send("Error al hacer la tirada.", ephemeral=True)

    async def handle_investigate_command(self, ctx, objetivo: str):
        """Handle the investigate command."""
        try:
            await ctx.defer()

            player = ctx.user.name
            room, path = self.get_player_location(player)

            if not room:
                await ctx.followup.send(
                    "No estás en ninguna sala de escape.", ephemeral=True
                )
                return

            # Use the take_path method to handle the investigation
            response = self.take_path(player, objetivo)
            await ctx.followup.send(response)

        except Exception as e:
            logger.error(f"Error in investigate command: {e}")
            await ctx.followup.send("Error al investigar.", ephemeral=True)

    async def handle_inventory_command(self, ctx):
        """Handle the inventory command."""
        try:
            await ctx.defer()

            player = ctx.user.name
            room, _ = self.get_player_location(player)

            if not room:
                await ctx.followup.send(
                    "No estás en ninguna sala de escape.", ephemeral=True
                )
                return

            inventory = self.get_inventory_dict(room)

            if not inventory:
                await ctx.followup.send("No tienes ningún objeto.", ephemeral=True)
                return

            embeds = []
            for item_name, item_desc in inventory.items():
                embed = discord.Embed(title=item_name, description=item_desc)
                embeds.append(embed)

            if embeds:
                view = PaginationView(embeds)
                await ctx.followup.send(embed=embeds[0], view=view, ephemeral=True)
            else:
                await ctx.followup.send("No tienes ningún objeto.", ephemeral=True)

        except Exception as e:
            logger.error(f"Error in inventory command: {e}")
            await ctx.followup.send("Error al mostrar inventario.", ephemeral=True)

    async def handle_equip_command(self, ctx, objeto: str):
        """Handle the equip command."""
        try:
            await ctx.defer()

            player = ctx.user.name
            response = self.equip(player, objeto)
            await ctx.followup.send(response)

        except Exception as e:
            logger.error(f"Error in equip command: {e}")
            await ctx.followup.send("Error al equipar objeto.", ephemeral=True)

    async def handle_combine_command(self, ctx, objeto1: str, objeto2: str):
        """Handle the combine command."""
        try:
            await ctx.defer()

            player = ctx.user.name
            room, _ = self.get_player_location(player)

            if not room:
                await ctx.followup.send(
                    "No estás en ninguna sala de escape.", ephemeral=True
                )
                return

            response = self.combine(objeto1, objeto2, room)
            await ctx.followup.send(response)

        except Exception as e:
            logger.error(f"Error in combine command: {e}")
            await ctx.followup.send("Error al combinar objetos.", ephemeral=True)

    async def handle_join_command(self, ctx):
        """Handle the join command."""
        try:
            await ctx.defer()

            player = ctx.user.name
            room, _ = self.get_player_location(player)

            if room:
                await ctx.followup.send(
                    "Aún no has escapado de la sala de huida en la que estás.",
                    ephemeral=True,
                )
                return

            channel_name = ctx.channel.name
            response = self.join_room(player, channel_name)
            await ctx.followup.send(response)

        except Exception as e:
            logger.error(f"Error in join command: {e}")
            await ctx.followup.send("Error al unirse a la sala.", ephemeral=True)
